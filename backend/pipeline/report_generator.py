"""
Report Generator for QueryCraft
Exports query results as CSV, Excel, or PDF — all in-memory, no disk writes.
"""
import csv
import io
import base64
from typing import Dict, List, Tuple, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Headless matplotlib configuration for thread safety
_HAS_MATPLOTLIB = False
try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    _HAS_MATPLOTLIB = True
except ImportError:
    pass

try:
    from pipeline.executor import detect_chart_type
except ImportError:
    try:
        from executor import detect_chart_type
    except ImportError:
        def detect_chart_type(columns: List[str]) -> str:
            return "table"

# PDF: try WeasyPrint first (requires GTK on Windows), fall back to reportlab
_PDF_BACKEND = None
try:
    import weasyprint
    _PDF_BACKEND = "weasyprint"
except Exception:
    pass

if _PDF_BACKEND is None:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Preformatted
        )
        _PDF_BACKEND = "reportlab"
    except ImportError:
        pass


# ── MIME types ────────────────────────────────────────────────────────────────
MIME_CSV   = "text/csv"
MIME_EXCEL = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MIME_PDF   = "application/pdf"


# ── CSV ───────────────────────────────────────────────────────────────────────
def export_csv(columns: List[str], rows: List[Dict]) -> bytes:
    """
    Export results as UTF-8 CSV bytes.

    Args:
        columns: Column names (header row)
        rows:    List of row dicts

    Returns:
        UTF-8 encoded CSV bytes
    """
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore",
                            lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


# ── Excel ─────────────────────────────────────────────────────────────────────
def export_excel(columns: List[str], rows: List[Dict]) -> bytes:
    """
    Export results as .xlsx bytes.

    Args:
        columns: Column names
        rows:    List of row dicts

    Returns:
        .xlsx file bytes
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QueryCraft Results"

    # Header style
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center")

    # Write header row
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align

    # Write data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = row.get(col_name)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns (cap at 50 chars)
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 54)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _generate_chart_image(columns: List[str], rows: List[Dict], chart_type: str) -> Optional[bytes]:
    """
    Generate a chart image of the specified chart_type ("bar" or "line") using Matplotlib and return PNG bytes.
    Returns None if generation is not possible or chart_type is invalid.
    """
    if not _HAS_MATPLOTLIB or not rows or not columns or not chart_type:
        return None

    chart_type = chart_type.lower().strip()
    if chart_type not in ("bar", "line"):
        return None

    # Identify X and Y columns
    columns_lower = [c.lower() for c in columns]
    x_col = None

    # Try to find a timestamp column for X-axis
    for i, col in enumerate(columns_lower):
        if 'timestamp' in col:
            x_col = columns[i]
            break

    # Fallback to first column as X-axis
    if not x_col and columns:
        x_col = columns[0]

    if not x_col:
        return None

    # Identify numeric Y columns (excluding the X column)
    y_cols = []
    # Limit data to first 30 rows for readable chart
    plot_data = rows[:30]

    for col in columns:
        if col == x_col:
            continue
        # Check if the column contains numeric values that we can plot
        is_numeric = False
        for r in plot_data:
            val = r.get(col)
            if val is not None:
                try:
                    float(val)
                    is_numeric = True
                    break
                except (ValueError, TypeError):
                    pass
        if is_numeric:
            y_cols.append(col)

    if not y_cols:
        return None

    try:
        fig, ax = plt.subplots(figsize=(8, 4))
        primary_color = "#1F4E79"

        # Extract X values
        x_vals = []
        for r in plot_data:
            val = r.get(x_col, "")
            # Shorten timestamps for visibility
            if isinstance(val, str) and len(val) > 19 and 'T' in val:
                x_vals.append(val.split('T')[1][:8])
            else:
                x_vals.append(str(val))

        if chart_type == "line":
            for y_col in y_cols:
                y_vals = []
                for r in plot_data:
                    try:
                        y_vals.append(float(r.get(y_col, 0)))
                    except (ValueError, TypeError):
                        y_vals.append(0.0)
                ax.plot(x_vals, y_vals, label=y_col, marker='o', linewidth=2, color=primary_color)
        elif chart_type == "bar":
            import numpy as np
            n_series = len(y_cols)
            x_indices = np.arange(len(plot_data))
            bar_width = 0.8 / n_series

            for idx, y_col in enumerate(y_cols):
                y_vals = []
                for r in plot_data:
                    try:
                        y_vals.append(float(r.get(y_col, 0)))
                    except (ValueError, TypeError):
                        y_vals.append(0.0)
                offset = (idx - (n_series - 1) / 2) * bar_width
                ax.bar(x_indices + offset, y_vals, bar_width, label=y_col, color=primary_color if idx == 0 else None)

            ax.set_xticks(x_indices)
            ax.set_xticklabels(x_vals)

        # Customize labels and style
        ax.set_title(f"{chart_type.capitalize()} Chart", fontsize=12, fontweight='bold', color='#333333')
        ax.set_xlabel(x_col, fontsize=10, fontweight='bold')
        if len(y_cols) == 1:
            ax.set_ylabel(y_cols[0], fontsize=10, fontweight='bold')

        ax.grid(True, linestyle='--', alpha=0.5, which='both', axis='y')
        ax.set_axisbelow(True)

        if len(x_vals) > 5:
            plt.xticks(rotation=45, ha='right')

        if len(y_cols) > 1:
            ax.legend(loc='best')

        plt.tight_layout()

        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=300)
        plt.close(fig)
        buf.seek(0)
        return buf.read()
    except Exception as e:
        print(f"[ReportGenerator] Failed to generate chart: {e}")
        return None


# ── PDF ───────────────────────────────────────────────────────────────────────
def export_pdf(columns: List[str], rows: List[Dict],
               query_text: str = "", sql: str = "",
               include_chart: bool = True,
               include_table: bool = True,
               chart_types: Optional[List[str]] = None,
               chart_type_override: Optional[str] = None) -> bytes:
    """
    Export results as PDF bytes.

    Uses WeasyPrint if GTK is available, otherwise falls back to reportlab.

    Args:
        columns:    Column names
        rows:       List of row dicts
        query_text: Original user query (shown in report header)
        sql:        Generated SQL (shown in report)
        include_chart: Whether to include chart in the PDF
        include_table: Whether to include data table in the PDF
        chart_types: List of chart types to include
        chart_type_override: Override for chart type (backwards compatibility)

    Returns:
        PDF file bytes
    """
    resolved_chart_types = []
    if include_chart:
        actual_chart_types = chart_types
        if actual_chart_types is None and chart_type_override is not None:
            actual_chart_types = [chart_type_override]

        if actual_chart_types:
            for ct in actual_chart_types:
                if not ct:
                    continue
                ct_lower = ct.lower().strip()
                if ct_lower == "auto":
                    detected = detect_chart_type(columns, rows)
                    if detected in ("bar", "line"):
                        resolved_chart_types.append(detected)
                else:
                    resolved_chart_types.append(ct_lower)
        else:
            detected = detect_chart_type(columns, rows)
            if detected in ("bar", "line"):
                resolved_chart_types.append(detected)

    if _PDF_BACKEND == "weasyprint":
        return _pdf_weasyprint(columns, rows, query_text, sql, resolved_chart_types, include_table)
    elif _PDF_BACKEND == "reportlab":
        return _pdf_reportlab(columns, rows, query_text, sql, resolved_chart_types, include_table)
    else:
        raise RuntimeError("No PDF backend available. Install reportlab: pip install reportlab")


def _pdf_weasyprint(columns, rows, query_text, sql,
                    resolved_chart_types: List[str],
                    include_table: bool = True) -> bytes:
    """Generate PDF using WeasyPrint (HTML → PDF)."""
    # Build HTML table rows
    header_cells = "".join(f"<th>{c}</th>" for c in columns)
    data_rows = ""
    for row in rows:
        cells = "".join(f"<td>{row.get(c, '')}</td>" for c in columns)
        data_rows += f"<tr>{cells}</tr>\n"

    chart_html = ""
    for ct in resolved_chart_types:
        chart_bytes = _generate_chart_image(columns, rows, ct)
        if chart_bytes:
            base64_data = base64.b64encode(chart_bytes).decode('utf-8')
            chart_html += f"""
            <div style="text-align: center; margin: 20px 0;">
                <img src="data:image/png;base64,{base64_data}" style="max-width: 100%; max-height: 350px; border: 1px solid #ddd; border-radius: 4px; padding: 5px;"/>
            </div>
            """

    table_html = ""
    if include_table:
        table_html = f"""
  <h3>Results ({len(rows)} rows)</h3>
  <table>
    <thead><tr>{header_cells}</tr></thead>
    <tbody>{data_rows}</tbody>
  </table>"""

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11px; margin: 20px; }}
  h1   {{ color: #1F4E79; font-size: 18px; }}
  h3   {{ color: #333; font-size: 13px; margin-top: 16px; }}
  pre  {{ background: #f4f4f4; padding: 8px; border-radius: 4px;
          font-size: 10px; white-space: pre-wrap; word-break: break-all; overflow-wrap: break-word; }}
  table {{ border-collapse: collapse; width: 100%; margin-top: 12px; table-layout: fixed; }}
  th, td {{ border: 1px solid #ddd; padding: 5px 8px; word-wrap: break-word; word-break: break-all; overflow-wrap: break-word; }}
  tr:nth-child(even) {{ background: #f9f9f9; }}
</style>
</head>
<body>
  <h1>QueryCraft Report</h1>
  <h3>Query</h3>
  <p>{query_text}</p>
  <h3>Generated SQL</h3>
  <pre>{sql}</pre>
  {chart_html}
  {table_html}
</body>
</html>"""

    return weasyprint.HTML(string=html).write_pdf()


def _pdf_reportlab(columns, rows, query_text, sql,
                    resolved_chart_types: List[str],
                    include_table: bool = True) -> bytes:
    """Generate PDF using reportlab (pure Python, no GTK needed)."""
    buf = io.BytesIO()

    # XML escape helper
    def _escape_xml(text: str) -> str:
        return str(text).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    # Use landscape if many columns
    pagesize = landscape(A4) if len(columns) > 6 else A4
    doc = SimpleDocTemplate(buf, pagesize=pagesize,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"],
                                 textColor=colors.HexColor("#1F4E79"), fontSize=16)
    h3_style    = ParagraphStyle("h3", parent=styles["Heading3"],
                                 textColor=colors.HexColor("#333333"), fontSize=11)
    body_style  = styles["BodyText"]
    
    # Code style with wordWrap enabled to prevent SQL overflow
    code_style  = ParagraphStyle("code", parent=styles["Code"],
                                 fontSize=8, leading=10,
                                 backColor=colors.HexColor("#F4F4F4"),
                                 wordWrap='CJK')

    # Table styles with wordWrap enabled to prevent text cell overflow
    table_header_style = ParagraphStyle(
        "TableHeader",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.white,
        alignment=1, # Center
        wordWrap='CJK'
    )
    table_cell_style = ParagraphStyle(
        "TableCell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7,
        leading=9,
        textColor=colors.HexColor("#333333"),
        wordWrap='CJK'
    )

    story = []

    # Title
    story.append(Paragraph("QueryCraft Report", title_style))
    story.append(Spacer(1, 0.3*cm))

    # Query text
    if query_text:
        story.append(Paragraph("Query", h3_style))
        story.append(Paragraph(query_text, body_style))
        story.append(Spacer(1, 0.2*cm))

    # SQL block
    if sql:
        story.append(Paragraph("Generated SQL", h3_style))
        # Escape HTML entities, replace newlines and multiple spaces for Paragraph rendering
        formatted_sql = _escape_xml(sql).replace('\n', '<br/>').replace('  ', '&nbsp;&nbsp;')
        story.append(Paragraph(formatted_sql, code_style))
        story.append(Spacer(1, 0.3*cm))

    # Chart images
    for ct in resolved_chart_types:
        chart_bytes = _generate_chart_image(columns, rows, ct)
        if chart_bytes:
            try:
                from reportlab.platypus import Image as RLImage
                img_flowable = RLImage(io.BytesIO(chart_bytes), width=16*cm, height=8*cm)
                story.append(img_flowable)
                story.append(Spacer(1, 0.4*cm))
            except Exception as e:
                print(f"[ReportGenerator] ReportLab image embedding failed: {e}")

    # Results table
    if include_table:
        story.append(Paragraph(f"Results ({len(rows)} rows)", h3_style))
        story.append(Spacer(1, 0.2*cm))

        if rows:
            # Build table data: header + rows (cap at 500 rows in PDF)
            display_rows = rows[:500]
            table_data = []
            
            # Header row
            table_data.append([Paragraph(_escape_xml(c), table_header_style) for c in columns])
            
            # Data rows
            for row in display_rows:
                row_cells = []
                for c in columns:
                    val = row.get(c, "")
                    row_cells.append(Paragraph(_escape_xml(val), table_cell_style))
                table_data.append(row_cells)

            # Column widths — distribute evenly
            page_w = pagesize[0] - 3*cm
            col_w  = page_w / len(columns)
            col_widths = [col_w] * len(columns)

            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(TableStyle([
                # Header
                ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("ALIGN",       (0, 0), (-1, 0), "CENTER"),
                # Data rows
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.white, colors.HexColor("#F9F9F9")]),
                ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#DDDDDD")),
                ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING",  (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(tbl)

            if len(rows) > 500:
                story.append(Spacer(1, 0.2*cm))
                story.append(Paragraph(
                    f"Note: PDF shows first 500 of {len(rows)} rows. "
                    "Download CSV or Excel for full dataset.",
                    body_style
                ))
        else:
            story.append(Paragraph("No results returned.", body_style))

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ── Entry point ───────────────────────────────────────────────────────────────
def generate_report(format: str, columns: List[str], rows: List[Dict],
                    query_text: str = "", sql: str = "",
                    include_chart: bool = True,
                    include_table: bool = True,
                    chart_types: Optional[List[str]] = None,
                    chart_type_override: Optional[str] = None) -> Tuple[bytes, str]:
    """
    Generate a report in the requested format.

    Args:
        format:     "csv", "excel", or "pdf"
        columns:    Column names
        rows:       List of row dicts
        query_text: Original user query
        sql:        Generated SQL
        include_chart: Whether to include chart in the PDF
        include_table: Whether to include data table in the PDF
        chart_types: List of chart types to include
        chart_type_override: Override for chart type (backwards compatibility)

    Returns:
        (file_bytes, mime_type)

    Raises:
        ValueError: If format is unsupported
    """
    fmt = format.lower().strip()

    if fmt == "csv":
        return export_csv(columns, rows), MIME_CSV

    elif fmt in ("excel", "xlsx"):
        return export_excel(columns, rows), MIME_EXCEL

    elif fmt == "pdf":
        return export_pdf(columns, rows, query_text, sql, include_chart, include_table, chart_types, chart_type_override), MIME_PDF

    else:
        raise ValueError(f"Unsupported format: '{format}'. Use 'csv', 'excel', or 'pdf'.")


# ── Self-test ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Testing Report Generator...")
    print(f"PDF backend: {_PDF_BACKEND}")
    print("=" * 80)

    # Sample data
    COLUMNS = ["cpu_num", "avg_busy_time", "system_name"]
    ROWS = [
        {"cpu_num": 0, "avg_busy_time": 1234567, "system_name": r"\KRONOS"},
        {"cpu_num": 1, "avg_busy_time": 987654,  "system_name": r"\KRONOS"},
        {"cpu_num": 2, "avg_busy_time": 1111111, "system_name": r"\KRONOS"},
    ]
    QUERY = "Show average CPU busy time per CPU"
    SQL   = "SELECT cpu_num, AVG(cpu_busy_time) AS avg_busy_time, system_name FROM macht413.cpu GROUP BY cpu_num, system_name LIMIT 10000"

    passed = 0
    total  = 3

    # ── CSV ──────────────────────────────────────────────────────────────────
    print("\n[Test 1] CSV Export")
    csv_bytes, mime = generate_report("csv", COLUMNS, ROWS)
    csv_text = csv_bytes.decode("utf-8")
    lines = csv_text.strip().split("\n")

    assert mime == MIME_CSV,                    f"Wrong MIME: {mime}"
    assert lines[0] == "cpu_num,avg_busy_time,system_name", f"Bad header: {lines[0]}"
    assert len(lines) == 4,                     f"Expected 4 lines, got {len(lines)}"
    assert "1234567" in csv_text,               "Missing data value"
    print(f"  MIME: {mime}")
    print(f"  Lines: {len(lines)} (header + {len(lines)-1} rows)")
    print(f"  Header: {lines[0]}")
    print("  ✓ PASSED")
    passed += 1

    # ── Excel ─────────────────────────────────────────────────────────────────
    print("\n[Test 2] Excel Export")
    xl_bytes, mime = generate_report("excel", COLUMNS, ROWS)

    assert mime == MIME_EXCEL,      f"Wrong MIME: {mime}"
    assert len(xl_bytes) > 0,       "Empty bytes"
    assert xl_bytes[:4] == b"PK\x03\x04", "Not a valid .xlsx (missing ZIP header)"

    # Verify content by re-reading with openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xl_bytes))
    ws = wb.active
    assert ws.cell(1, 1).value == "cpu_num",    "Header cell wrong"
    assert ws.cell(2, 1).value == 0,            "Data cell wrong"
    assert ws.cell(1, 1).font.bold,             "Header not bold"

    print(f"  MIME: {mime}")
    print(f"  Size: {len(xl_bytes):,} bytes")
    print(f"  Rows in sheet: {ws.max_row} (header + {ws.max_row - 1} data)")
    print("  ✓ PASSED")
    passed += 1

    # ── PDF ───────────────────────────────────────────────────────────────────
    print("\n[Test 3] PDF Export")
    pdf_bytes, mime = generate_report("pdf", COLUMNS, ROWS, QUERY, SQL)

    assert mime == MIME_PDF,            f"Wrong MIME: {mime}"
    assert len(pdf_bytes) > 0,          "Empty bytes"
    assert pdf_bytes[:4] == b"%PDF",    f"Not a valid PDF (got {pdf_bytes[:4]})"

    print(f"  MIME: {mime}")
    print(f"  Backend: {_PDF_BACKEND}")
    print(f"  Size: {len(pdf_bytes):,} bytes")
    print("  ✓ PASSED")
    passed += 1

    # ── Bad format ────────────────────────────────────────────────────────────
    print("\n[Test 4] Unsupported Format")
    try:
        generate_report("docx", COLUMNS, ROWS)
        print("  ✗ Should have raised ValueError")
    except ValueError as e:
        print(f"  Correctly raised ValueError: {e}")
        print("  ✓ PASSED")

    print("\n" + "=" * 80)
    print(f"✓ {passed}/{total} export tests passed!")
    print("=" * 80)
