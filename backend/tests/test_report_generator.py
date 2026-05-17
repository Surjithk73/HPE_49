"""
Comprehensive tests for Report Generator (CSV, Excel, PDF)
"""
import csv
import io
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
from pipeline.report_generator import (
    generate_report, export_csv, export_excel, export_pdf,
    MIME_CSV, MIME_EXCEL, MIME_PDF, _PDF_BACKEND
)

# ── Shared sample data ────────────────────────────────────────────────────────
COLUMNS = ["cpu_num", "avg_busy_time", "system_name"]
ROWS = [
    {"cpu_num": 0, "avg_busy_time": 1234567, "system_name": r"\KRONOS"},
    {"cpu_num": 1, "avg_busy_time": 987654,  "system_name": r"\KRONOS"},
    {"cpu_num": 2, "avg_busy_time": 1111111, "system_name": r"\KRONOS"},
]
QUERY = "Show average CPU busy time per CPU"
SQL   = "SELECT cpu_num, AVG(cpu_busy_time) FROM macht413.cpu GROUP BY cpu_num LIMIT 10000"


def test_csv():
    print("\n" + "=" * 80)
    print("TEST 1: CSV Export")
    print("=" * 80)

    checks = []

    # Basic export
    data, mime = generate_report("csv", COLUMNS, ROWS)
    text  = data.decode("utf-8")
    lines = text.strip().split("\n")

    # MIME type
    ok = mime == MIME_CSV
    print(f"  MIME type correct: {ok}  ({mime})")
    checks.append(ok)

    # Header row
    ok = lines[0] == "cpu_num,avg_busy_time,system_name"
    print(f"  Header row correct: {ok}  ({lines[0]})")
    checks.append(ok)

    # Row count
    ok = len(lines) == 4   # 1 header + 3 data
    print(f"  Row count correct: {ok}  ({len(lines)} lines)")
    checks.append(ok)

    # Data values present
    ok = "1234567" in text and r"\KRONOS" in text
    print(f"  Data values present: {ok}")
    checks.append(ok)

    # Valid CSV (re-parse)
    reader = list(csv.DictReader(io.StringIO(text)))
    ok = len(reader) == 3 and reader[0]["cpu_num"] == "0"
    print(f"  Re-parseable as CSV: {ok}")
    checks.append(ok)

    # Empty rows
    empty_bytes, _ = generate_report("csv", COLUMNS, [])
    empty_lines = empty_bytes.decode("utf-8").strip().split("\n")
    ok = len(empty_lines) == 1   # header only
    print(f"  Empty rows → header only: {ok}")
    checks.append(ok)

    # In-memory (no disk write)
    ok = isinstance(data, bytes)
    print(f"  Returns bytes (in-memory): {ok}")
    checks.append(ok)

    passed = sum(checks)
    print(f"\nCSV: {passed}/{len(checks)} checks passed")
    return all(checks)


def test_excel():
    print("\n" + "=" * 80)
    print("TEST 2: Excel Export")
    print("=" * 80)

    checks = []

    data, mime = generate_report("excel", COLUMNS, ROWS)

    # MIME type
    ok = mime == MIME_EXCEL
    print(f"  MIME type correct: {ok}")
    checks.append(ok)

    # Valid .xlsx magic bytes (ZIP PK header)
    ok = data[:4] == b"PK\x03\x04"
    print(f"  Valid .xlsx magic bytes: {ok}")
    checks.append(ok)

    # Non-empty
    ok = len(data) > 1000
    print(f"  Non-empty ({len(data):,} bytes): {ok}")
    checks.append(ok)

    # Re-open with openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active

    # Header row content
    ok = ws.cell(1, 1).value == "cpu_num"
    print(f"  Header cell correct: {ok}  ({ws.cell(1,1).value})")
    checks.append(ok)

    # Header bold
    ok = ws.cell(1, 1).font.bold is True
    print(f"  Header is bold: {ok}")
    checks.append(ok)

    # Data rows
    ok = ws.max_row == 4   # 1 header + 3 data
    print(f"  Row count correct: {ok}  ({ws.max_row} rows)")
    checks.append(ok)

    # Data value
    ok = ws.cell(2, 1).value == 0
    print(f"  First data cell correct: {ok}  ({ws.cell(2,1).value})")
    checks.append(ok)

    # In-memory
    ok = isinstance(data, bytes)
    print(f"  Returns bytes (in-memory): {ok}")
    checks.append(ok)

    passed = sum(checks)
    print(f"\nExcel: {passed}/{len(checks)} checks passed")
    return all(checks)


def test_pdf():
    print("\n" + "=" * 80)
    print("TEST 3: PDF Export")
    print("=" * 80)
    print(f"  PDF backend: {_PDF_BACKEND}")

    checks = []

    data, mime = generate_report("pdf", COLUMNS, ROWS, QUERY, SQL)

    # MIME type
    ok = mime == MIME_PDF
    print(f"  MIME type correct: {ok}")
    checks.append(ok)

    # Starts with %PDF
    ok = data[:4] == b"%PDF"
    print(f"  Starts with %PDF: {ok}  ({data[:8]})")
    checks.append(ok)

    # Non-empty (at least 1 KB)
    ok = len(data) > 1000
    print(f"  Non-empty ({len(data):,} bytes): {ok}")
    checks.append(ok)

    # In-memory
    ok = isinstance(data, bytes)
    print(f"  Returns bytes (in-memory): {ok}")
    checks.append(ok)

    # Empty rows still produces valid PDF
    empty_data, _ = generate_report("pdf", COLUMNS, [], QUERY, SQL)
    ok = empty_data[:4] == b"%PDF"
    print(f"  Empty rows → valid PDF: {ok}")
    checks.append(ok)

    passed = sum(checks)
    print(f"\nPDF: {passed}/{len(checks)} checks passed")
    return all(checks)


def test_generate_report_routing():
    print("\n" + "=" * 80)
    print("TEST 4: generate_report() Routing & MIME Types")
    print("=" * 80)

    checks = []

    # csv
    _, mime = generate_report("csv", COLUMNS, ROWS)
    ok = mime == MIME_CSV
    print(f"  csv  → {mime}: {ok}")
    checks.append(ok)

    # excel
    _, mime = generate_report("excel", COLUMNS, ROWS)
    ok = mime == MIME_EXCEL
    print(f"  excel → {mime}: {ok}")
    checks.append(ok)

    # xlsx alias
    _, mime = generate_report("xlsx", COLUMNS, ROWS)
    ok = mime == MIME_EXCEL
    print(f"  xlsx  → {mime}: {ok}")
    checks.append(ok)

    # pdf
    _, mime = generate_report("pdf", COLUMNS, ROWS)
    ok = mime == MIME_PDF
    print(f"  pdf  → {mime}: {ok}")
    checks.append(ok)

    # unsupported format
    try:
        generate_report("docx", COLUMNS, ROWS)
        checks.append(False)
        print("  docx → should have raised ValueError: ✗")
    except ValueError:
        checks.append(True)
        print("  docx → ValueError raised: ✓")

    passed = sum(checks)
    print(f"\nRouting: {passed}/{len(checks)} checks passed")
    return all(checks)


def test_large_dataset():
    """Verify all formats handle 1000-row dataset without error."""
    print("\n" + "=" * 80)
    print("TEST 5: Large Dataset (1000 rows)")
    print("=" * 80)

    big_rows = [{"cpu_num": i, "avg_busy_time": i * 1000, "system_name": r"\KRONOS"}
                for i in range(1000)]

    checks = []
    for fmt in ("csv", "excel", "pdf"):
        try:
            data, _ = generate_report(fmt, COLUMNS, big_rows)
            ok = len(data) > 0
            print(f"  {fmt:6s}: {len(data):,} bytes  ✓")
            checks.append(ok)
        except Exception as e:
            print(f"  {fmt:6s}: ✗ {e}")
            checks.append(False)

    passed = sum(checks)
    print(f"\nLarge Dataset: {passed}/{len(checks)} formats handled")
    return all(checks)


def run_all_tests():
    print("\n" + "=" * 80)
    print("REPORT GENERATOR — COMPREHENSIVE TEST SUITE")
    print("=" * 80)

    results = [
        ("CSV Export",          test_csv()),
        ("Excel Export",        test_excel()),
        ("PDF Export",          test_pdf()),
        ("Routing & MIME",      test_generate_report_routing()),
        ("Large Dataset",       test_large_dataset()),
    ]

    print("\n" + "=" * 80)
    print("TEST SUMMARY")
    print("=" * 80)
    for name, ok in results:
        print(f"{name:30s}: {'✓ PASSED' if ok else '✗ FAILED'}")

    all_passed = all(r[1] for r in results)
    print("\n" + "=" * 80)
    print("✓ ALL REPORT GENERATOR TESTS PASSED" if all_passed else "✗ SOME TESTS FAILED")
    print("=" * 80)
    return all_passed


if __name__ == "__main__":
    success = run_all_tests()
    sys.exit(0 if success else 1)
